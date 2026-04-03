from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count
from django.utils import timezone
from django.http import HttpResponse
from .models import Student, Employee, Fee, Attendance, UserProfile, Receipt, ReceiptItem
from .forms import StudentForm, EmployeeForm, FeeForm, AttendanceForm, ReceiptForm, ReceiptItemFormSet


def login_view(request):
    """Handle user login."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        role = request.POST.get('role')
        
        # Try to find user by email or username
        user = None
        from django.contrib.auth.models import User
        
        # First try to authenticate with email as username
        user = authenticate(request, username=email, password=password)
        
        # If that fails, try to find user by email field
        if user is None:
            try:
                user_obj = User.objects.get(email=email)
                user = authenticate(request, username=user_obj.username, password=password)
            except User.DoesNotExist:
                pass
        
        if user is not None:
            # Check role
            try:
                profile = user.userprofile
                if profile.role == role:
                    login(request, user)
                    messages.success(request, 'Login successful!')
                    return redirect('dashboard')
                else:
                    messages.error(request, 'Invalid role selected.')
            except User.DoesNotExist:
                messages.error(request, 'User not found.')
            except UserProfile.DoesNotExist:
                messages.error(request, 'User profile not found.')
        else:
            messages.error(request, 'Invalid email or password.')
    
    return render(request, 'login.html')


def logout_view(request):
    """Handle user logout."""
    logout(request)
    messages.info(request, 'You have been logged out.')
    return redirect('login')


@login_required
def dashboard(request):
    """Display dashboard with statistics."""
    # Get statistics
    total_students = Student.objects.count()
    total_employees = Employee.objects.count()
    
    # Calculate total revenue
    revenue = Fee.objects.aggregate(total=Sum('amount_paid'))['total'] or 0
    
    # Get today's attendance count
    today = timezone.now().date()
    attendance_today = Attendance.objects.filter(date=today).count()
    
    context = {
        'total_students': total_students,
        'total_employees': total_employees,
        'revenue': revenue,
        'attendance_today': attendance_today,
        'current_date': today.strftime('%A, %B %d, %Y'),
    }
    return render(request, 'dashboard.html', context)


@login_required
def students(request):
    """Handle student management."""
    search_query = request.GET.get('search', '')
    
    if search_query:
        students_list = Student.objects.filter(name__icontains=search_query)
    else:
        students_list = Student.objects.all()
    
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student added successfully!')
            return redirect('students')
    else:
        form = StudentForm()
    
    context = {
        'students': students_list,
        'form': form,
        'search_query': search_query,
    }
    return render(request, 'students.html', context)


@login_required
def attendance(request):
    """Handle attendance marking."""
    if request.method == 'POST':
        form = AttendanceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Attendance recorded successfully!')
            return redirect('attendance')
    else:
        form = AttendanceForm()
    
    context = {
        'form': form,
    }
    return render(request, 'attendance.html', context)


@login_required
def fees(request):
    """Handle fee payments."""
    if request.method == 'POST':
        form = FeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Payment recorded successfully!')
            return redirect('fees')
    else:
        form = FeeForm()
    
    context = {
        'form': form,
    }
    return render(request, 'fees.html', context)


@login_required
def payroll(request):
    """Handle payroll management."""
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Employee added successfully!')
            return redirect('payroll')
    else:
        form = EmployeeForm()
    
    context = {
        'form': form,
    }
    return render(request, 'payroll.html', context)


@login_required
def generate_payslip(request):
    """Generate payslip for an employee."""
    employee = None
    
    if request.method == 'POST':
        emp_name = request.POST.get('emp_name', '')
        if emp_name:
            try:
                employee = Employee.objects.get(name__icontains=emp_name)
                messages.success(request, f'Payslip generated for {employee.name}')
            except Employee.DoesNotExist:
                messages.error(request, 'Employee not found.')
            except Employee.MultipleObjectsReturned:
                employee = Employee.objects.filter(name__icontains=emp_name).first()
    
    context = {
        'employee': employee,
    }
    return render(request, 'payslip.html', context)


# Receipt Views
@login_required
def receipts(request):
    """Display all receipts."""
    search_query = request.GET.get('search', '')
    receipts_list = Receipt.objects.all().select_related('student')
    
    if search_query:
        receipts_list = receipts_list.filter(
            receipt_number__icontains=search_query
        ) | receipts_list.filter(
            student__name__icontains=search_query
        )
    
    context = {
        'receipts': receipts_list,
        'search_query': search_query,
    }
    return render(request, 'receipts.html', context)


@login_required
def create_receipt(request):
    """Create a new receipt with multiple items."""
    if request.method == 'POST':
        form = ReceiptForm(request.POST)
        formset = ReceiptItemFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            receipt = form.save()
            items = formset.save(commit=False)
            
            for item in items:
                item.receipt = receipt
                item.save()
            
            # Calculate and update total
            receipt.calculate_total()
            receipt.save()
            
            messages.success(request, f'Receipt #{receipt.receipt_number} created successfully!')
            return redirect('receipts')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ReceiptForm()
        formset = ReceiptItemFormSet()
    
    context = {
        'form': form,
        'formset': formset,
    }
    return render(request, 'create_receipt.html', context)


@login_required
def receipt_detail(request, receipt_id):
    """View receipt details."""
    receipt = get_object_or_404(Receipt, id=receipt_id)
    context = {
        'receipt': receipt,
    }
    return render(request, 'receipt_detail.html', context)


# Excel Export View
@login_required
def export_debtors(request):
    """Export students with arrears to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Debtors Report"
        
        # Headers
        headers = ['Student Name', 'Class', 'Total Fee', 'Total Paid', 'Balance/Arrears']
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Get students with fees
        students = Student.objects.all()
        for student in students:
            fees = Fee.objects.filter(student=student)
            total_fee = fees.aggregate(Sum('total_fee'))['total_fee__sum'] or 0
            total_paid = fees.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
            balance = total_fee - total_paid
            
            # Only include if there's a balance (arrears)
            if balance > 0:
                ws.append([
                    student.name,
                    student.student_class or 'N/A',
                    float(total_fee),
                    float(total_paid),
                    float(balance)
                ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=debtors_report.xlsx'
        
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl. Please install it: pip install openpyxl')
        return redirect('fees')
